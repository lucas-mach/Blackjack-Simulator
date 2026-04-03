from bj import Game, PlayerHand, DealerHand, Deck, Card
import pandas as pd
from openpyxl import load_workbook
import copy

# Ordered from highest TCC to lowest (matches determine_strategy priority)
_TCC_KEY_THRESHOLDS = [
    ('tcc_8plus',       8),
    ('tcc_7',           7),
    ('tcc_6',           6),
    ('tcc_5',           5),
    ('tcc_4',           4),
    ('tcc_3',           3),
    ('tcc_2',           2),
    ('tcc_0_1',         0),
    ('tcc_neg1',       -1),
    ('tcc_neg2',       -2),
    ('tcc_under_neg2', -3),
]

_TCC_BASE_PATH = 'strategies/strategy_tcc_0_1.xlsx'


def _resolve_strategy_key(true_count: float, strategy_overrides: dict, tcc_key_to_path: dict) -> str:
    """Resolve the strategy file path for a given TCC when user-defined override rows exist.

    Rules:
    - Empty dict  -> use base strategy (tcc_0_1)
    - Has entries -> prefer highest defined threshold that is <= true_count (i.e. 'at or below').
                     If none are <= true_count, use the lowest defined threshold (floor fallback).
    """
    if not strategy_overrides:           # {} or None -> base strategy
        return tcc_key_to_path.get('tcc_0_1', _TCC_BASE_PATH)

    defined = [(k, t) for k, t in _TCC_KEY_THRESHOLDS if k in strategy_overrides]
    if not defined:
        return tcc_key_to_path.get('tcc_0_1', _TCC_BASE_PATH)

    # Prefer highest threshold that is still <= current count
    below = [(k, t) for k, t in defined if t <= true_count]
    if below:
        best_key = max(below, key=lambda x: x[1])[0]   # highest threshold <= count
    else:
        best_key = min(defined, key=lambda x: x[1])[0]  # floor: lowest defined

    return tcc_key_to_path.get(best_key, _TCC_BASE_PATH)


class AutoGame:
    
    def count_cards(items):
        """Return Hi-Lo count for a collection which may contain hands or raw cards."""
        total = 0
        # If we got a single hand, wrap it in a list
        if not isinstance(items, list):
            items = [items]
            
        for item in items:
            if hasattr(item, 'cards'):
                # It's a Hand object
                item_cards = item.cards
            elif hasattr(item, 'rank'):
                # It's a Card object
                item_cards = [item]
            elif isinstance(item, list):
                # It's already a list of cards
                item_cards = item
            else:
                continue
                
            for c in item_cards:
                r = str(c.rank)
                if r in ('2', '3', '4', '5', '6'):
                    total += 1
                elif r in ('10', 'J', 'Q', 'K', 'A'):
                    total -= 1
        return total

    def determine_strategy(true_card_count: int) -> str:
        """
        Returns the appropriate blackjack strategy file based on the true card count (TCC).
        
        Strategy files incorporate I18 deviations from the deviation chart:
        
        TCC >= 8:  13v10=SURR
        TCC >= 7:  14v9=SURR, 15v8=SURR, 10v10=DBL
        TCC >= 6:  10-10 vs 5=SPLIT
        TCC >= 5:  10-10 vs 6=SPLIT, 14vA=SURR, 9v7=DBL, 10vA=DBL
        TCC >= 4:  16v8=SURR
        TCC >= 3:  14v10=SURR, 12v2=STAND
        TCC >= 2:  15v9=SURR, 15vA=SURR, 9v2=DBL, 11vA=DBL(already base), 12v3=STAND
        TCC >= 0:  15v10=SURR, 16vA=SURR
        TCC >= -1: 16v9=SURR, 16v10=SURR
        TCC >= -1: 12v6=STAND (base=S, else Hit), 13v2=STAND (base=S, else Hit)
        TCC >= -2: 12v5=STAND (base=S, else Hit), 13v3=STAND (base=S, else Hit)
        TCC >= -2: 12v4=STAND (base=S, else Hit at TCC < 0)
        """
        if true_card_count >= 8:
            return "strategies/strategy_tcc_8plus.xlsx"
        elif true_card_count >= 7:
            return "strategies/strategy_tcc_7.xlsx"
        elif true_card_count >= 6:
            return "strategies/strategy_tcc_6.xlsx"
        elif true_card_count >= 5:
            return "strategies/strategy_tcc_5.xlsx"
        elif true_card_count >= 4:
            return "strategies/strategy_tcc_4.xlsx"
        elif true_card_count >= 3:
            return "strategies/strategy_tcc_3.xlsx"
        elif true_card_count >= 2:
            return "strategies/strategy_tcc_2.xlsx"
        elif true_card_count >= 0:
            return "strategies/strategy_tcc_0_1.xlsx"
        elif true_card_count >= -1:
            return "strategies/strategy_tcc_neg1.xlsx"
        elif true_card_count >= -2:
            return "strategies/strategy_tcc_neg2.xlsx"
        else:
            return "strategies/strategy_tcc_under_neg2.xlsx"

            
    def determine_bet_multiple(true_count: int) -> int:
        """
        Bet spread based on the Kelly Criterion and professional Hi-Lo counting strategy.
        Uses a 1-12 unit spread, the most common among professional counters.
        
        True Count | Bet (units) | Player Edge (approx)
        -----------|-------------|---------------------
            < 1   |      1      |   House edge
            1   |      2      |   ~0%  (break even)
            2   |      4      |   ~0.5%
            3   |      6      |   ~1.0%
            4   |      8      |   ~1.5%
            5   |      10     |   ~2.0%
            6+   |      12     |   ~2.5%+
        -----------|-------------|---------------------
        Negative counts: table minimum (sit out if allowed)
        """
        BET_RAMP = {
            1: 2,
            2: 4,
            3: 6,
            4: 8,
            5: 10,
        }
        
        if true_count <= 0:
            return 1   # Minimum bet — house has the edge
        if true_count >= 6:
            return 12  # Maximum bet — Kelly-optimal cap
        
        return BET_RAMP.get(true_count, 1)

    class Strategy:
        def __init__(self, path: str):
            self.hard_df, self.soft_df, self.split_df = self.read_exact_ranges(path)

        def read_exact_ranges(self, path: str):
            wb = load_workbook(path, data_only=True)
            ws = wb.active

            def get_range(ws, ref):
                data = ws[ref]
                rows = [[cell.value for cell in row] for row in data]
                return pd.DataFrame(rows)

            hard = get_range(ws, "B3:K20")
            soft = get_range(ws, "B23:K32")
            split = get_range(ws, "B35:K44")
            return hard, soft, split

        def get_action(self, player_hand: PlayerHand, dealer_value, splitallowed=True) -> str:
            if player_hand.can_split() and splitallowed:
                row = player_hand.cards[0].get_value() - 2
                col = dealer_value - 2
                if self.split_df.iat[row, col] == 'Y':
                    return 'v'
            hand_value = player_hand.get_value()
            if player_hand.has_soft_ace():
                row = hand_value - 12
                col = dealer_value - 2
                action = self.soft_df.iat[row, col]
                if action == "DS":
                    return 'd' if player_hand.num_cards() == 2 else 's'
                else:
                    if action == "D":
                        return 'd' if player_hand.num_cards() == 2 else 'h'
                    return action.lower()
            else:
                row = hand_value - 4
                col = dealer_value - 2
                action = self.hard_df.iat[row, col]
                if action == "R":
                    action = "H"  
                if action == "RS":
                    action = "S" 
                if action == "D" and player_hand.num_cards() != 2:
                    action = "H"
                return action.lower()

        def apply_overrides(self, overrides: dict):
            """Return a new Strategy with specific cells patched.
            overrides = {hard: [[row, col, val], ...], soft: [...], split: [...]}
            """
            new_strat = copy.copy(self)
            new_strat.hard_df = self.hard_df.copy()
            new_strat.soft_df = self.soft_df.copy()
            new_strat.split_df = self.split_df.copy()
            for r, c, v in overrides.get('hard', []):
                new_strat.hard_df.iat[int(r), int(c)] = str(v).upper()
            for r, c, v in overrides.get('soft', []):
                new_strat.soft_df.iat[int(r), int(c)] = str(v).upper()
            for r, c, v in overrides.get('split', []):
                new_strat.split_df.iat[int(r), int(c)] = str(v).upper()
            return new_strat

    def auto_play_loop(bet_amount=1, num_games=100, balance=1000, num_decks=8,
                       input_func=input, output_func=print, return_as_json=False,
                       rules=None,
                       bet_ramp=None,
                       strategy_overrides=None,
                       insurance_threshold=None,
                       use_base_strategy_only=False):
        if num_games <= 0:
            if return_as_json:
                return {"error": "Number of games must be at least 1.", "results": [], "logs": [],
                        "final_balance": balance, "total_profit": 0}
            else:
                output_func("Number of games must be at least 1.")
                return

        deck = Deck(num_decks)
        deck.shuffle()
        total_profit = 0
        shoe_profit = 0
        card_count = 0
        MAX_SPLITS = 4
        # Data storage
        results = []
        logs = []

        def local_output(*args):
            logs.append(" ".join(map(str, args)))
            output_func(*args)

        # legacy text file removed; use csv when not returning json
        #if not return_as_json:
        #    open('results.txt', 'w').close()

        all_paths = [
            "strategies/strategy_tcc_8plus.xlsx",
            "strategies/strategy_tcc_7.xlsx",
            "strategies/strategy_tcc_6.xlsx",
            "strategies/strategy_tcc_5.xlsx",
            "strategies/strategy_tcc_4.xlsx",
            "strategies/strategy_tcc_3.xlsx",
            "strategies/strategy_tcc_2.xlsx",
            "strategies/strategy_tcc_0_1.xlsx",
            "strategies/strategy_tcc_neg1.xlsx",
            "strategies/strategy_tcc_neg2.xlsx",
            "strategies/strategy_tcc_under_neg2.xlsx",
        ]
        strategies = {path: AutoGame.Strategy(path) for path in all_paths}

        # Apply any user strategy overrides per TCC key
        TCC_KEY_TO_PATH = {
            'tcc_8plus':       'strategies/strategy_tcc_8plus.xlsx',
            'tcc_7':           'strategies/strategy_tcc_7.xlsx',
            'tcc_6':           'strategies/strategy_tcc_6.xlsx',
            'tcc_5':           'strategies/strategy_tcc_5.xlsx',
            'tcc_4':           'strategies/strategy_tcc_4.xlsx',
            'tcc_3':           'strategies/strategy_tcc_3.xlsx',
            'tcc_2':           'strategies/strategy_tcc_2.xlsx',
            'tcc_0_1':         'strategies/strategy_tcc_0_1.xlsx',
            'tcc_neg1':        'strategies/strategy_tcc_neg1.xlsx',
            'tcc_neg2':        'strategies/strategy_tcc_neg2.xlsx',
            'tcc_under_neg2':  'strategies/strategy_tcc_under_neg2.xlsx',
        }
        if strategy_overrides:
            for tcc_key, overrides in strategy_overrides.items():
                path = TCC_KEY_TO_PATH.get(tcc_key)
                if path and path in strategies:
                    strategies[path] = strategies[path].apply_overrides(overrides)

        strategy_path = AutoGame.determine_strategy(0)
        strategy = strategies[strategy_path]

        for i in range(num_games):
            if balance <= 0:
                output_func("Out of money! Balance is 0.")
                break
            
            if i > 0 and i % 1000 == 0:
                output_func(f"Progress: {i}/{num_games} games completed. Current Balance: {balance}")

            true_card_count = card_count / (len(deck.cards)/52) if len(deck.cards) > 0 else 0

            # Custom bet ramp: list of 7 multipliers [le0, 1, 2, 3, 4, 5, ge6]
            if bet_ramp and len(bet_ramp) >= 7:
                tcc_int = int(true_card_count)
                if tcc_int <= 0:
                    mult = bet_ramp[0]
                elif tcc_int >= 6:
                    mult = bet_ramp[6]
                else:
                    mult = bet_ramp[min(tcc_int, 5)]
            else:
                mult = AutoGame.determine_bet_multiple(true_card_count)
            base_bet = bet_amount * mult
            #base_bet = max(base_bet, bet_amount)   # <----------TESTING ONLY


            # Bet max if money is less than bet
            modified_bet_amount = min(base_bet, balance)
            
            if bet_amount == 0:
                modified_bet_amount = 0
            
            if modified_bet_amount == 0 and bet_amount > 0:
                output_func("Time to leave the table, count is too low.")
                deck = Deck(num_decks=num_decks)
                deck.shuffle()
                card_count = 0
                shoe_profit = 0
                continue
            
            # Strategy selection: respect user-defined override rows if present
            if strategy_overrides is not None:
                # User has configured the strategy editor (even if empty = base fallback)
                new_strategy_path = _resolve_strategy_key(
                    true_card_count, strategy_overrides, TCC_KEY_TO_PATH
                )
            elif use_base_strategy_only:
                new_strategy_path = AutoGame.determine_strategy(0)  # always tcc_0_1
            else:
                new_strategy_path = AutoGame.determine_strategy(true_card_count)
            if new_strategy_path != strategy_path:
                strategy_path = new_strategy_path
                strategy = strategies[strategy_path]

            game = deck.new_game()
            # Pass balance MINUS initial bet to played_hand for splits/doubles
            round_result = AutoGame.played_hand(game, modified_bet_amount, balance - modified_bet_amount, strategy, input_func=input_func, output_func=output_func, MAX_SPLITS=MAX_SPLITS)

            # Insurance side-bet: applied after hand is played, dealer cards already known
            ins_profit = 0
            if (insurance_threshold is not None and
                    true_card_count >= insurance_threshold and
                    len(game.dealer_hand.cards) > 0 and
                    game.dealer_hand.cards[0].rank == 'A'):
                ins_bet = modified_bet_amount * 0.5
                if game.dealer_hand.blackjack():
                    ins_profit = ins_bet * 2   # 2:1 payout (net +ins_bet)
                else:
                    ins_profit = -ins_bet

            profit = Game.interpret_result(round_result) + ins_profit
            balance += profit
            total_profit += profit
            shoe_profit += profit
            
            card_count += Game.interpret_card_count(round_result)
            
            # Compute actual total bet placed this hand (handles splits & doubles)
            if isinstance(round_result, list):
                actual_bet = sum(r[1] for r in round_result)
            else:
                actual_bet = round_result[1]
            
            true_card_count_log = card_count / (len(deck.cards)/52) if len(deck.cards) > 0 else 0
            results.append({
                "hand": i+1,
                "balance": balance,
                "card_count": card_count,
                "true_count": true_card_count_log,
                "bet": actual_bet,
            })
            game.end_game()
            
            if len(deck.cards) < (52 * num_decks * 0.25):
                    output_func("Reshuffling deck. Shoe Profit: ", shoe_profit)
                    deck = Deck(num_decks=num_decks)
                    deck.shuffle()
                    card_count = 0
                    shoe_profit = 0

        # convert accumulated rows into a pandas DataFrame and save csv
        results_df = pd.DataFrame(results)
        csv_saved = False
        try:
            results_df.to_csv('results.csv', index=False)
            csv_saved = True
        except Exception:
            # if writing fails, still return dataframe
            pass

        if return_as_json:
            return {
                "num_games": num_games,
                "results": results,
                "logs": logs,
                "final_balance": balance,
                "total_profit": total_profit
            }
        else:
            # Auto-generate the simulation graph after a console/script run
            if csv_saved:
                try:
                    from graph import SimGraph
                    SimGraph().generate()
                except Exception as e:
                    output_func(f"[SimGraph] Could not generate graph: {e}")
            return results_df

    def played_hand(game, bet_amount, balance, strategy, input_func=input, output_func=print, MAX_SPLITS=4):
        game.deal_initial()
        dealer_hand = game.dealer_hand
        
        # Initial Blackjack check
        if game.player_hands[0].blackjack():
            if dealer_hand.blackjack():
                return ("P", bet_amount, AutoGame.count_cards([game.player_hands[0], dealer_hand]))
            return ("W!", bet_amount, AutoGame.count_cards([game.player_hands[0], dealer_hand]))
        if dealer_hand.blackjack():
            return ("L", bet_amount, AutoGame.count_cards([game.player_hands[0], dealer_hand]))

        bets = [bet_amount]
        played_indices = []
        
        # Iteratively play all player hands (supports multiple splits)
        i = 0
        while i < len(game.player_hands):
            hand = game.player_hands[i]
            current_bet = bets[i]
            
            while not hand.is_busted():
                if hand.get_value() >= 21:
                    break
                    
                # Action based on strategy
                action = strategy.get_action(hand, dealer_hand.get_card_shown_value(), splitallowed=(len(game.player_hands) < MAX_SPLITS))
                
                if action == 'h':
                    game.player_hit(handnum=i)
                elif action == 's':
                    break
                elif action == 'd' and hand.num_cards() == 2 and balance >= current_bet:
                    game.player_hit(handnum=i)
                    bets[i] *= 2
                    break
                elif action == 'v' and hand.can_split() and balance >= current_bet:
                    balance -= current_bet
                    game.deal_split(i)
                    # When we split, the current hand 'i' gets a new card, and a new hand is at i+1
                    bets.insert(i + 1, current_bet)
                    # Continue playing the current hand i
                    continue
                else:
                    # Fallback to stand if something is wrong
                    break
            i += 1
            
        # Dealer plays if any hand isn't busted
        any_active = any(not h.is_busted() for h in game.player_hands)
        if any_active:
            game.dealer_play(Auto=True)
            
        # Collect results
        results = []
        total_count = 0
        for idx, h in enumerate(game.player_hands):
            res = game.get_winner(idx)
            cnt = 0
            if idx == 0:
                # Only count dealer cards once
                total_cards = []
                for ph in game.player_hands:
                    total_cards.extend(ph.cards)
                total_cards.extend(dealer_hand.cards)
                total_count = AutoGame.count_cards(total_cards)
                cnt = total_count
            results.append((res, bets[idx], cnt))
            
        if len(results) == 1:
            return results[0]
        return results
